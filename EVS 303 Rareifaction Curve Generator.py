import pandas as pd
import os
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import copy
import random
import pandas as pd
from openpyxl import load_workbook
import time
os.chdir(os.path.dirname(os.path.abspath(__file__)))

workbookDirectory = "Soil Invertebrate Count Data.xlsx"
#workbookDirectory = 'Master Code Folder/Python Projects/EVS 303 Rareifaction Curve Generator/Soil Invertebrate Count Data.xlsx'
RawDataop = load_workbook(workbookDirectory)
RawData = pd.ExcelFile(workbookDirectory)
ws = RawDataop["Sheet3"]
# Strange = str(input("Where does your range start? "))
Strange = "G1"
Enrange = "X31"
# Enrange = str(input("Where does your range end? "))
# Question = str(input("Does your selection include labels? y/n"))
range1 = ws[Strange:Enrange]
#print(range)
i = 0
dataConstructor = list()
for cell in range1:
    newList = []
    #print(cell)
    dataConstructor.append(newList)
    i+=1
    for x in cell:
        value = str(x.value)
        newList.append(value)
        #print(newList)


iterationNumber = len(dataConstructor[0])

x=0
ListofColumns = []
newList = []
for iteration in range(iterationNumber):
    newList=[]
    for lists in dataConstructor:
        newList.append(lists[x])
        
    x+=1
    ListofColumns.append(newList)

noHeaders = copy.deepcopy(ListofColumns)
for orgs in noHeaders:
    del orgs[0]
    

x = 0
for i in noHeaders:
    try:
        for k,j in zip(i,range(len(noHeaders[x]))):
            noHeaders[x][j] = int(k)
    except IndexError:
        pass
    x+=1

df = pd.DataFrame(noHeaders)

totalAbundance = list(df.sum(0))
totalIndividuals = list(df.sum(1))

new_df = df.transpose()
#print(new_df.values)

df1 = df.values.tolist()
newdf = new_df.values.tolist()
#print(newdf)



def uniqueGrabber(newdf):
    x=0
    uniqueSpecies = []
    for rows in newdf:
        counter = 0
        #print(rows)
        y=0
        for number in rows:
            if number != 0:
                checker = []
                for i in range(x):
                    checker.append(newdf[i][y])
                if sum(checker) == 0:
                    counter+=1
            y+=1
        x+=1
        uniqueSpecies.append(counter)
    return uniqueSpecies
def speciesCounter(newdf):
    numberOfSpecies = []
    for rows in newdf:
        x=0
        for number in rows:
            if number != 0:
                x+=1
        numberOfSpecies.append(x)
    return numberOfSpecies


shuffleNumber = 10
uniqueSpeciesCount = []
cumNumberOfSpecies = []

random_lists=[]
random_transposed = []
y1=0
for shuffle in range(shuffleNumber):
    for lists in df1:
        y2=0
        newList = random.sample(lists, len(lists))
        for numbers in newList:
            lists[y2] = numbers
            y2+=1
            
    newDF = pd.DataFrame(df1)
    newdfdf = newDF.transpose()
    lister = newdfdf.values.tolist()
    
    random_lists.append(lister)
            
for r in random_lists:
    uniqueSpeciesCount.append(uniqueGrabber(r))

            
        
    
#for k in random_lists:
    #print(k)
#     Udf = pd.DataFrame(r)
#     Udf.transpose()
#     newUdf = Udf.values.tolist()
#     uniqueSpeciesCount.append(uniqueGrabber(newUdf))
#     y1+=1
    
for unique in uniqueSpeciesCount:
    cum = 0
    cums = []
    for u in unique:
        cum+=u
        cums.append(cum)
    cumNumberOfSpecies.append(cums)
    
#print(uniqueSpeciesCount)
#print(cumNumberOfSpecies)


wb = load_workbook("Soil Invertebrate Count Data.xlsx")
ws = wb["Sheet3"]


col = 32
for random in cumNumberOfSpecies:
    row = 2
    for r in random:
        ws.cell(column=col, row=row, value=r)
        row+=1
    col+=1
    
    
col = 32
for random in uniqueSpeciesCount:
    row = 36
    for r in random:
        ws.cell(column=col, row=row, value=r)
        row+=1
    col+=1
    
wb.save("Soil Invertebrate Count Data.xlsx")

print("""
░█████╗░░█████╗░██╗░░░░░░█████╗░██╗░░░██╗██╗░░░░░░█████╗░████████╗██╗░█████╗░███╗░░██╗
██╔══██╗██╔══██╗██║░░░░░██╔══██╗██║░░░██║██║░░░░░██╔══██╗╚══██╔══╝██║██╔══██╗████╗░██║
██║░░╚═╝███████║██║░░░░░██║░░╚═╝██║░░░██║██║░░░░░███████║░░░██║░░░██║██║░░██║██╔██╗██║
██║░░██╗██╔══██║██║░░░░░██║░░██╗██║░░░██║██║░░░░░██╔══██║░░░██║░░░██║██║░░██║██║╚████║
╚█████╔╝██║░░██║███████╗╚█████╔╝╚██████╔╝███████╗██║░░██║░░░██║░░░██║╚█████╔╝██║░╚███║
░╚════╝░╚═╝░░╚═╝╚══════╝░╚════╝░░╚═════╝░╚══════╝╚═╝░░╚═╝░░░╚═╝░░░╚═╝░╚════╝░╚═╝░░╚══╝

░█████╗░░█████╗░███╗░░░███╗██████╗░██╗░░░░░███████╗████████╗███████╗
██╔══██╗██╔══██╗████╗░████║██╔══██╗██║░░░░░██╔════╝╚══██╔══╝██╔════╝
██║░░╚═╝██║░░██║██╔████╔██║██████╔╝██║░░░░░█████╗░░░░░██║░░░█████╗░░
██║░░██╗██║░░██║██║╚██╔╝██║██╔═══╝░██║░░░░░██╔══╝░░░░░██║░░░██╔══╝░░
╚█████╔╝╚█████╔╝██║░╚═╝░██║██║░░░░░███████╗███████╗░░░██║░░░███████╗
░╚════╝░░╚════╝░╚═╝░░░░░╚═╝╚═╝░░░░░╚══════╝╚══════╝░░░╚═╝░░░╚══════╝""")
time.sleep(10)

